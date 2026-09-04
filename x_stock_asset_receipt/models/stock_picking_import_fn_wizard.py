import base64
import io
import openpyxl

from odoo import _, fields, models
from odoo.exceptions import UserError


class StockPickingImportFnWizard(models.TransientModel):
    _name = 'stock.picking.import.fn.wizard'
    _description = 'Wizard Import Data FN Excel'

    picking_id = fields.Many2one(
        'stock.picking',
        string='Receipt',
        required=True,
        readonly=True,
    )
    file = fields.Binary(
        string='File Excel (.xlsx)',
        required=True,
    )
    filename = fields.Char(string='Nama File')
    state = fields.Selection(
        [('upload', 'Upload'), ('result', 'Result')],
        default='upload',
    )
    warning_message = fields.Html(
        string='Hasil Import / Warning',
        readonly=True,
    )
    success_count = fields.Integer(
        string='Jumlah Sukses',
        default=0,
        readonly=True,
    )
    failed_count = fields.Integer(
        string='Jumlah Gagal / Peringatan',
        default=0,
        readonly=True,
    )

    def action_import(self):
        self.ensure_one()
        if not self.file:
            raise UserError(_('Silakan pilih file Excel terlebih dahulu.'))

        try:
            file_data = base64.b64decode(self.file)
            wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
            ws = wb.active
        except Exception as e:
            raise UserError(_('File tidak valid atau rusak: %s') % str(e))

        # Baca header di baris 1
        headers = {}
        for col_idx, cell in enumerate(ws[1], start=1):
            val = str(cell.value or '').strip().lower()
            if val:
                headers[val] = col_idx

        def find_col(aliases):
            for alias in aliases:
                if alias.lower() in headers:
                    return headers[alias.lower()]
            return None

        col_line_no = find_col([
            'line no', 'line no.', 'line', 'item no', 'item no.', 'item',
            'no', 'no.', 'nomor line', 'nomor item'
        ])
        col_product = find_col(['product', 'produk', 'nama produk', 'item name'])
        col_chassis = find_col(['chassis number', 'chassis', 'no rangka', 'no. rangka', 'nomor rangka'])
        col_engine = find_col(['engine number', 'engine', 'no mesin', 'no. mesin', 'nomor mesin'])
        col_plate = find_col(['initial license plate', 'license plate', 'plat nomor', 'nopol', 'plat'])
        col_warna = find_col(['warna', 'color'])
        col_tahun = find_col(['tahun', 'year'])
        col_fn = find_col(['fleet number', 'fn', 'asset number', 'serial number'])

        if not col_line_no:
            raise UserError(_("Kolom 'Line No' / 'Item No' tidak ditemukan di dalam file Excel."))

        if not col_fn:
            raise UserError(_("Kolom 'Fleet Number' (FN) tidak ditemukan di dalam file Excel."))

        picking = self.picking_id
        active_moves = picking.move_ids.filtered(lambda m: m.state != 'cancel')
        moves_by_line_no = {
            idx: move
            for idx, move in enumerate(active_moves, start=1)
            if move.product_id.is_vehicle
        }
        vehicle_moves = picking.move_ids.filtered(
            lambda m: m.product_id.is_vehicle and m.state != 'cancel'
        )
        has_fn = any(
            (l.lot_id or l.lot_name)
            for m in vehicle_moves
            for l in m.move_line_ids
        )
        if not vehicle_moves or not has_fn:
            raise UserError(
                _('Belum ada line penerimaan dengan Fleet Number (FN) pada receipt ini. '
                  'Silakan jalankan "Mass Generate FN" terlebih dahulu.')
            )

        # Mapping FN ke (line_no, move, line) untuk validasi pencocokan
        lines_by_fn = {}
        for move_idx, move in moves_by_line_no.items():
            for line in move.move_line_ids:
                fn_name = (line.lot_id.name or line.lot_name or '').strip()
                if fn_name:
                    lines_by_fn[fn_name] = (move_idx, move, line)

        warnings = []
        success_count = 0
        failed_count = 0
        processed_fns = set()

        # Cache master warna dan tahun dalam lowercase
        colors = {c.name.strip().lower(): c for c in self.env['vehicle.color'].search([]) if c.name}
        years = {y.name.strip().lower(): y for y in self.env['vehicle.year'].search([]) if y.name}

        for row_idx in range(2, ws.max_row + 1):
            def get_val(col_idx):
                if not col_idx:
                    return ''
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is None:
                    return ''
                if isinstance(val, float) and val.is_integer():
                    return str(int(val)).strip()
                return str(val).strip()

            line_no_raw = get_val(col_line_no)
            fn_val = get_val(col_fn)
            chassis_val = get_val(col_chassis)
            engine_val = get_val(col_engine)
            plate_val = get_val(col_plate)
            warna_val = get_val(col_warna)
            tahun_val = get_val(col_tahun)

            # Lewati baris jika semua kolom kosong
            if not any([line_no_raw, fn_val, chassis_val, engine_val, plate_val, warna_val, tahun_val]):
                continue

            if not line_no_raw:
                warnings.append(
                    _("Baris Excel %d: 'Line No' kosong.") % row_idx
                )
                failed_count += 1
                continue

            try:
                line_no = int(float(line_no_raw))
            except (ValueError, TypeError):
                warnings.append(
                    _("Baris Excel %d: 'Line No' (%s) tidak valid.") % (row_idx, line_no_raw)
                )
                failed_count += 1
                continue

            if not fn_val:
                warnings.append(
                    _("Line item %s (FN: -): Fleet Number kosong di baris Excel %d.")
                    % (line_no, row_idx)
                )
                failed_count += 1
                continue

            if fn_val in processed_fns:
                warnings.append(
                    _("Line item %s (FN: %s): Fleet Number terduplikasi di file Excel pada baris %d.")
                    % (line_no, fn_val, row_idx)
                )
                failed_count += 1
                continue
            processed_fns.add(fn_val)

            # Validasi keberadaan Line No di receipt
            target_move = moves_by_line_no.get(line_no)
            if not target_move:
                warnings.append(
                    _("Line item %s (FN: %s): Line No tidak ditemukan pada receipt ini (total line receipt: %d).")
                    % (line_no, fn_val, len(active_moves))
                )
                failed_count += 1
                continue

            # Validasi apakah FN terdaftar di receipt ini
            fn_info = lines_by_fn.get(fn_val)
            if not fn_info:
                warnings.append(
                    _("Line item %s (FN: %s): Fleet Number '%s' tidak ditemukan pada receipt ini.")
                    % (line_no, fn_val, fn_val)
                )
                failed_count += 1
                continue

            actual_line_no, actual_move, target_line = fn_info

            # Validasi kecocokan Line No dan Fleet Number
            if actual_line_no != line_no:
                warnings.append(
                    _("Line item %s (FN: %s): Fleet Number tidak cocok. FN '%s' seharusnya berada di Line %d pada receipt.")
                    % (line_no, fn_val, fn_val, actual_line_no)
                )
                failed_count += 1
                continue

            # Resolusi warna bila diisi: cek lowercase, buat baru jika belum ada (uppercase huruf pertama)
            color_record = False
            if warna_val:
                color_key = warna_val.strip().lower()
                color_record = colors.get(color_key)
                if not color_record:
                    color_record = self.env['vehicle.color'].search([('name', '=ilike', warna_val.strip())], limit=1)
                if not color_record:
                    formatted_color = warna_val.strip().capitalize()
                    color_record = self.env['vehicle.color'].create({'name': formatted_color})
                colors[color_key] = color_record

            # Resolusi tahun bila diisi: cek lowercase, buat baru jika belum ada (uppercase huruf pertama)
            year_record = False
            if tahun_val:
                year_key = tahun_val.strip().lower()
                year_record = years.get(year_key)
                if not year_record:
                    year_record = self.env['vehicle.year'].search([('name', '=ilike', tahun_val.strip())], limit=1)
                if not year_record:
                    formatted_year = tahun_val.strip().capitalize()
                    year_record = self.env['vehicle.year'].create({'name': formatted_year})
                years[year_key] = year_record

            # Baris valid: update data line receipt (dan otomatis tersinkron ke lot_id)
            vals_to_write = {}
            if chassis_val:
                vals_to_write['chassis_number'] = chassis_val
            if engine_val:
                vals_to_write['engine_number'] = engine_val
            if plate_val:
                vals_to_write['initial_license_plate'] = plate_val
            if warna_val and color_record:
                vals_to_write['vehicle_color_id'] = color_record.id
            if tahun_val and year_record:
                vals_to_write['vehicle_year_id'] = year_record.id

            if vals_to_write:
                target_line.write(vals_to_write)
            success_count += 1

        self.success_count = success_count
        self.failed_count = failed_count

        if warnings:
            warning_items = "".join([f"<li>{w}</li>" for w in warnings])
            msg = f"""
            <div class="alert alert-warning" role="alert" style="margin-bottom: 12px;">
                <strong>Import selesai dengan catatan:</strong><br/>
                Sebanyak <b>{success_count}</b> baris berhasil diperbarui. Terdapat <b>{failed_count}</b> baris tidak cocok / bermasalah:
            </div>
            <ul style="max-height: 220px; overflow-y: auto; padding-left: 20px; color: #856404;">
                {warning_items}
            </ul>
            """
            self.warning_message = msg
            self.state = 'result'

            # Catat juga ke chatter dokumen receipt
            chatter_msg = (
                f"<b>Hasil Import Data FN Excel:</b><br/>"
                f"- Berhasil: {success_count} baris<br/>"
                f"- Peringatan / Tidak Cocok ({failed_count} baris):<br/>"
                + "<br/>".join([f"• {w}" for w in warnings])
            )
            picking.message_post(body=chatter_msg)

            return {
                'type': 'ir.actions.act_window',
                'name': _('Hasil Import FN Excel'),
                'res_model': self._name,
                'res_id': self.id,
                'view_mode': 'form',
                'target': 'new',
            }

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Import Sukses'),
                'message': _('Semua data (%s unit) berhasil diimpor.') % success_count,
                'type': 'success',
                'sticky': False,
                'next': {'type': 'ir.actions.act_window_close'},
            },
        }
