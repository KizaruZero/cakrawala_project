import base64
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from odoo import api, fields, models, _
from odoo.exceptions import UserError


class StockPicking(models.Model):
    _inherit = 'stock.picking'

    rental_type_id = fields.Many2one(
        'vehicle.substatus',
        string='Rental Type',
        domain=[('is_rental_type', '=', True)],
        ondelete='restrict',
        tracking=True,
        help='Sub-status flagged as Rental Type in Master Sub Status. '
             'The value picked here is applied as Fleet Sub-Status when the asset is registered.',
    )

    is_asset_registered = fields.Boolean(
        string="Asset Registered",
        compute='_compute_is_asset_registered',
        store=True,
        copy=False,
    )
    has_vehicle_product = fields.Boolean(
        string="Has Vehicle Product",
        compute='_compute_has_vehicle_product',
    )
    is_from_sales_order = fields.Boolean(
        string="Is From Sales Order",
        compute='_compute_is_from_sales_order',
    )

    def _compute_is_from_sales_order(self):
        for picking in self:
            picking.is_from_sales_order = hasattr(picking, 'sale_id') and bool(picking.sale_id)

    is_bastk_linked = fields.Boolean(
        string="Is BASTK Linked",
        compute='_compute_is_bastk_linked',
    )

    def _compute_is_bastk_linked(self):
        for picking in self:
            picking.is_bastk_linked = hasattr(picking, 'bastk_id') and bool(picking.bastk_id)

    @api.depends(
        'state',
        'picking_type_code',
        'move_line_ids.lot_id',
        'move_ids.product_id.is_vehicle',
    )
    def _compute_is_asset_registered(self):
        FleetVehicle = self.env['fleet.vehicle']
        for picking in self:
            if picking.picking_type_code != 'incoming' or picking.state != 'done':
                picking.is_asset_registered = False
                continue

            lot_lines = picking.move_line_ids.filtered(
                lambda ml: ml.lot_id and ml.product_id.is_vehicle
            )
            if lot_lines:
                all_registered = all(
                    FleetVehicle.search_count([('asset_number', '=', ml.lot_id.name)]) > 0
                    for ml in lot_lines
                )
                picking.is_asset_registered = all_registered
                continue

            picking.is_asset_registered = False

    @api.depends('move_ids.product_id.is_vehicle', 'move_ids.state')
    def _compute_has_vehicle_product(self):
        for picking in self:
            active_moves = picking.move_ids.filtered(
                lambda m: m.state != 'cancel' and m.product_id
            )
            picking.has_vehicle_product = any(
                move.product_id.is_vehicle for move in active_moves
            )

    def button_validate(self):
        """Override Validate: validasi mandatory fields per unit (move_line level)."""
        for picking in self:
            if picking.picking_type_code != 'incoming':
                continue

            missing = []

            # Rental Type hanya relevan kalau ada produk fleet (product.is_vehicle)
            # dan GR tidak terhubung dengan BASTK (is_bastk_linked == False).
            has_done_vehicle = any(
                m.product_id.is_vehicle and m.quantity > 0
                for m in picking.move_ids.filtered(lambda m: m.state != 'cancel')
            )
            if has_done_vehicle and not picking.is_bastk_linked and not picking.rental_type_id:
                missing.append('Rental Type (header GR)')

            for move in picking.move_ids:
                if move.state in ('done', 'cancel'):
                    continue
                if move.product_id.tracking != 'serial':
                    continue

                # Pada partial receipt, lewati move yang kuantitas terimanya 0 (akan jadi backorder)
                if move.quantity <= 0:
                    continue

                is_vehicle = move.product_id.is_vehicle
                product_name = move.product_id.display_name

                if not move.move_line_ids:
                    missing.append(
                        'Detail Operations (produk: %s) — klik tombol Detail untuk mengisi data unit'
                        % product_name
                    )
                    continue

                if move.quantity > move.product_uom_qty:
                    missing.append(
                        'Kuantitas Done (%s) melebihi Demand (%s) untuk produk %s'
                        % (move.quantity, move.product_uom_qty, product_name)
                    )

                for idx, line in enumerate(move.move_line_ids, start=1):
                    unit_label = '%s (unit %d)' % (product_name, idx)

                    if line.quantity >= 1.0:
                        if not line.lot_id and not (line.lot_name or '').strip():
                            missing.append('Serial Number — %s' % unit_label)

                        if is_vehicle:
                            if not (line.initial_license_plate or '').strip():
                                missing.append('Initial License Plate — %s' % unit_label)
                            if not (line.chassis_number or '').strip():
                                missing.append('Chassis Number — %s' % unit_label)
                            if not (line.engine_number or '').strip():
                                missing.append('Engine Number — %s' % unit_label)
                            if not line.vehicle_color_id:
                                missing.append('Warna — %s' % unit_label)
                            if not line.vehicle_year_id:
                                missing.append('Tahun — %s' % unit_label)

            if missing:
                raise UserError(
                    _('Unable to validate GR. Please fill in the following fields:\n- %s')
                    % '\n- '.join(missing)
                )

        return super().button_validate()

    def _default_fleet_vehicle_state_for_gr(self):
        """Prefer is_first_destination; else Non-Leased; else standard Fleet Registered; else any state."""
        VehicleState = self.env['fleet.vehicle.state']
        state = VehicleState.search([('is_first_destination', '=', True)], limit=1)
        if not state:
            state = VehicleState.search([('name', '=', 'Non-Leased')], limit=1)
        if state:
            return state.id
        ref = self.env.ref('fleet.fleet_vehicle_state_registered', raise_if_not_found=False)
        if ref:
            return ref.id
        fallback = VehicleState.search([], limit=1, order='sequence, id')
        return fallback.id if fallback else False

    def _fleet_substatus_from_rental_type(self):
        """The GR rental type IS a vehicle.substatus record now — nothing to map."""
        self.ensure_one()
        return self.rental_type_id

    def action_register_asset_detail(self):
        """
        Buat fleet.vehicle untuk setiap unit (move_line dengan lot_id),
        lalu buka list view dari semua kendaraan yang dibuat.
        """
        self.ensure_one()

        default_state_id = self._default_fleet_vehicle_state_for_gr()

        vehicle_ids = []
        vehicle_lines = self.move_line_ids.filtered(
            lambda ml: ml.lot_id and ml.product_id.is_vehicle and ml.quantity >= 1.0
        )

        for line in vehicle_lines:
            existing = self.env['fleet.vehicle'].search(
                [('asset_number', '=', line.lot_id.name)], limit=1
            )
            if existing:
                vehicle_ids.append(existing.id)
                continue

            product = line.product_id
            model = product.fleet_model_id
            if not model:
                model = self.env['fleet.vehicle.model'].search([('name', '=', product.name)], limit=1)
                if not model:
                    brand = self.env['fleet.vehicle.model.brand'].search([('name', '=', 'Other')], limit=1)
                    if not brand:
                        brand = self.env['fleet.vehicle.model.brand'].create({'name': 'Other'})
                    
                    model = self.env['fleet.vehicle.model'].create({
                        'name': product.name,
                        'brand_id': brand.id,
                    })

            fleet_sub = self._fleet_substatus_from_rental_type()
            vehicle_vals = {
                'model_id': model.id,
                'asset_number': line.lot_id.name,
                'chassis_number': line.chassis_number or line.lot_id.chassis_number or '',
                'engine_number': line.engine_number or line.lot_id.engine_number or '',
                'initial_license_plate': line.initial_license_plate or line.lot_id.initial_license_plate or '',
                'fleet_sub_status_id': fleet_sub.id if fleet_sub else False,
                'state_id': default_state_id,
                'model_year': line.vehicle_year_id.name if line.vehicle_year_id else '',
                'color': line.vehicle_color_id.name if line.vehicle_color_id else '',
            }
            vehicle = self.env['fleet.vehicle'].create(vehicle_vals)
            vehicle_ids.append(vehicle.id)

            lot_vals = {}
            if line.vehicle_year_id:
                lot_vals['vehicle_year_id'] = line.vehicle_year_id.id
            if line.vehicle_color_id:
                lot_vals['vehicle_color_id'] = line.vehicle_color_id.id
            if lot_vals:
                line.lot_id.with_context(skip_sync_fleet=True).write(lot_vals)

        self._compute_is_asset_registered()

        if not vehicle_ids:
            raise UserError(
                _('No vehicle serial numbers found. Please generate Serial Numbers for each '
                  'vehicle unit in the Detailed Operations before registering.')
            )

        if len(vehicle_ids) == 1:
            return {
                'name': _('Fleet Registration'),
                'type': 'ir.actions.act_window',
                'res_model': 'fleet.vehicle',
                'view_mode': 'form',
                'res_id': vehicle_ids[0],
                'target': 'current',
                'context': dict(self.env.context, active_id=vehicle_ids[0], active_ids=[vehicle_ids[0]]),
            }

        return {
            'name': _('Fleet Registration'),
            'type': 'ir.actions.act_window',
            'res_model': 'fleet.vehicle',
            'view_mode': 'list,form',
            'domain': [('id', 'in', vehicle_ids)],
            'target': 'current',
            'context': dict(self.env.context, active_id=False, active_ids=vehicle_ids),
        }

    def action_mass_generate_fn(self):
        for picking in self:
            picking.move_ids.action_mass_generate_fn()
        return True

    def action_export_fn_excel(self):
        """Export data line receipt yang sudah punya FN ke file Excel (.xlsx)."""
        self.ensure_one()
        active_moves = self.move_ids.filtered(lambda m: m.state != 'cancel')
        move_to_line_no = {
            move.id: idx
            for idx, move in enumerate(active_moves, start=1)
        }
        vehicle_moves = active_moves.filtered(lambda m: m.product_id.is_vehicle)

        has_fn = any(
            (l.lot_id or l.lot_name)
            for m in vehicle_moves
            for l in m.move_line_ids
        )
        if not vehicle_moves or not has_fn:
            raise UserError(
                _('Belum ada line penerimaan dengan Fleet Number (FN). '
                  'Silakan jalankan "Mass Generate FN" terlebih dahulu.')
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Receipt FN Details"

        headers = [
            "Line No",
            "Product",
            "Chassis Number",
            "Engine Number",
            "Initial License Plate",
            "Warna",
            "Tahun",
            "Fleet Number",
        ]
        ws.append(headers)

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9'),
        )

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row_idx = 1
        for move in vehicle_moves:
            line_no = move_to_line_no[move.id]
            move_vehicle_lines = move.move_line_ids.filtered(
                lambda l: l.lot_id or l.lot_name
            )
            for line in move_vehicle_lines:
                fn = line.lot_id.name or line.lot_name or ''
                product_name = line.product_id.display_name or line.product_id.name or ''
                chassis = line.chassis_number or ''
                engine = line.engine_number or ''
                plate = line.initial_license_plate or ''
                warna = line.vehicle_color_id.name if line.vehicle_color_id else ''
                tahun = line.vehicle_year_id.name if line.vehicle_year_id else ''

                row = [
                    line_no,
                    product_name,
                    chassis,
                    engine,
                    plate,
                    warna,
                    tahun,
                    fn,
                ]
                ws.append(row)
                row_idx += 1
                for col_idx in range(1, len(headers) + 1):
                    c = ws.cell(row=row_idx, column=col_idx)
                    c.border = thin_border
                    if col_idx in (1, 6, 7, 8):
                        c.alignment = Alignment(horizontal="center", vertical="center")

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

        # ----------------------------------------------------
        # Sheet 2: Referensi Warna & Tahun + Petunjuk/Keterangan
        # ----------------------------------------------------
        ws_ref = wb.create_sheet(title="Referensi Data")

        # Judul & Keterangan Panduan
        ws_ref.merge_cells("A1:E1")
        ws_ref["A1"] = "PETUNJUK & REFERENSI MASTER DATA"
        ws_ref["A1"].font = Font(name="Calibri", size=11, bold=True, color="1F4E78")

        ws_ref.merge_cells("A2:E2")
        ws_ref["A2"] = "1. Anda dapat mengisi kolom 'Warna' dan 'Tahun' pada sheet 'Receipt FN Details' mengacu pada tabel referensi di bawah."
        ws_ref["A2"].font = Font(name="Calibri", size=10, color="495057")

        ws_ref.merge_cells("A3:E3")
        ws_ref["A3"] = (
            "2. CATATAN OTOMATISASI: Apabila warna dan/atau tahun yang Anda input BELUM ADA / TIDAK MATCH "
            "dengan daftar di bawah, sistem Odoo akan OTOMATIS MEMBUAT (GENERATE) master data warna dan/atau tahun baru tersebut saat file di-import."
        )
        ws_ref["A3"].font = Font(name="Calibri", size=10, bold=True, color="B25900")

        ws_ref.merge_cells("A4:E4")
        ws_ref["A4"] = "3. Penulisan nama warna dan tahun tidak sensitif huruf besar/kecil (sistem akan otomatis memformat huruf kapital di awal kata)."
        ws_ref["A4"].font = Font(name="Calibri", size=10, color="495057")

        # Header Tabel Referensi di Baris 6
        ref_headers = {
            1: ("No", "center"),
            2: ("Referensi Warna (Terdaftar)", "left"),
            3: ("", "center"),
            4: ("No", "center"),
            5: ("Referensi Tahun (Terdaftar)", "center"),
        }
        for col_idx, (header_text, align_h) in ref_headers.items():
            if col_idx == 3:
                continue
            c = ws_ref.cell(row=6, column=col_idx, value=header_text)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal=align_h, vertical="center")

        colors = self.env['vehicle.color'].search([], order='name asc')
        years = self.env['vehicle.year'].search([], order='name desc')
        color_list = [c.name for c in colors if c.name]
        year_list = [y.name for y in years if y.name]

        max_ref_rows = max(len(color_list), len(year_list), 1)
        for i in range(max_ref_rows):
            r_idx = 7 + i
            if i < len(color_list):
                c_no = ws_ref.cell(row=r_idx, column=1, value=i + 1)
                c_no.border = thin_border
                c_no.alignment = Alignment(horizontal="center", vertical="center")

                c_val = ws_ref.cell(row=r_idx, column=2, value=color_list[i])
                c_val.border = thin_border
                c_val.alignment = Alignment(horizontal="left", vertical="center")

            if i < len(year_list):
                y_no = ws_ref.cell(row=r_idx, column=4, value=i + 1)
                y_no.border = thin_border
                y_no.alignment = Alignment(horizontal="center", vertical="center")

                y_val = ws_ref.cell(row=r_idx, column=5, value=year_list[i])
                y_val.border = thin_border
                y_val.alignment = Alignment(horizontal="center", vertical="center")

        max_color_len = max([len(str(c)) for c in color_list] or [20])
        ws_ref.column_dimensions['A'].width = 8
        ws_ref.column_dimensions['B'].width = max(max_color_len + 6, 28)
        ws_ref.column_dimensions['C'].width = 4
        ws_ref.column_dimensions['D'].width = 8
        ws_ref.column_dimensions['E'].width = 24

        # Pastikan active sheet saat pertama dibuka adalah sheet utama
        wb.active = ws

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"FN_Receipt_{self.name.replace('/', '_')}.xlsx"
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': base64.b64encode(output.getvalue()),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    def action_open_import_fn_wizard(self):
        """Buka wizard untuk import data kendaraan dari file Excel."""
        self.ensure_one()
        if self.state in ('done', 'cancel'):
            raise UserError(_('Tidak dapat mengimpor data pada penerimaan yang sudah selesai atau dibatalkan.'))
        return {
            'type': 'ir.actions.act_window',
            'name': _('Import Data FN Excel'),
            'res_model': 'stock.picking.import.fn.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'default_picking_id': self.id,
            },
        }

